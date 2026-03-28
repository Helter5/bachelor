"""
Styling utilities for PDF and Excel exports
"""
from typing import List, Tuple
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import TableStyle
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .font_manager import font_manager


class ColorPalette:
    """Color definitions for consistent styling"""

    # Dark colors
    DARK_GRAY = colors.HexColor('#1f2937')
    MEDIUM_GRAY = colors.HexColor('#6b7280')
    LIGHT_GRAY = colors.HexColor('#f3f4f6')
    VERY_LIGHT_GRAY = colors.HexColor('#f9fafb')

    # Brand colors
    PRIMARY_BLUE = colors.HexColor('#3b82f6')

    # Medal colors
    GOLD = colors.HexColor('#fef3c7')
    SILVER = colors.HexColor('#e5e7eb')
    BRONZE = colors.HexColor('#fed7aa')

    # Chart colors
    CHART_COLORS = ['#FFD700', '#C0C0C0', '#CD7F32', '#4A90E2', '#50C878']

    # Basic colors
    WHITE = colors.white
    WHITESMOKE = colors.whitesmoke
    GREY = colors.grey


class PDFStyleFactory:
    """Factory for creating PDF table styles"""

    @staticmethod
    def create_header_style(
        col_range: Tuple[int, int] = (0, -1),
        row: int = 0,
        bg_color=None,
        text_color=None,
        font_size: int = 10,
        bold: bool = True
    ) -> List:
        """
        Create header style for table

        Args:
            col_range: Column range tuple (start, end)
            row: Row number
            bg_color: Background color
            text_color: Text color
            font_size: Font size
            bold: Use bold font

        Returns:
            List of style commands
        """
        bg_color = bg_color or ColorPalette.DARK_GRAY
        text_color = text_color or ColorPalette.WHITESMOKE

        return [
            ('BACKGROUND', (col_range[0], row), (col_range[1], row), bg_color),
            ('TEXTCOLOR', (col_range[0], row), (col_range[1], row), text_color),
            ('ALIGN', (col_range[0], row), (col_range[1], row), 'CENTER'),
            ('FONTNAME', (col_range[0], row), (col_range[1], row),
             font_manager.bold_font if bold else font_manager.default_font),
            ('FONTSIZE', (col_range[0], row), (col_range[1], row), font_size),
            ('BOTTOMPADDING', (col_range[0], row), (col_range[1], row), 12),
        ]

    @staticmethod
    def create_body_style(
        col_range: Tuple[int, int] = (0, -1),
        row_start: int = 1,
        row_end: int = -1,
        font_size: int = 9,
        alternating_colors: bool = True
    ) -> List:
        """
        Create body style for table

        Args:
            col_range: Column range tuple
            row_start: Start row
            row_end: End row
            font_size: Font size
            alternating_colors: Use alternating row colors

        Returns:
            List of style commands
        """
        styles = [
            ('FONTNAME', (col_range[0], row_start), (col_range[1], row_end),
             font_manager.default_font),
            ('FONTSIZE', (col_range[0], row_start), (col_range[1], row_end), font_size),
        ]

        if alternating_colors:
            styles.append(
                ('ROWBACKGROUNDS', (col_range[0], row_start), (col_range[1], row_end),
                 [ColorPalette.WHITE, ColorPalette.LIGHT_GRAY])
            )

        return styles

    @staticmethod
    def create_grid_style(
        col_range: Tuple[int, int] = (0, -1),
        row_range: Tuple[int, int] = (0, -1),
        line_width: float = 0.5,
        color=None
    ) -> List:
        """
        Create grid style

        Args:
            col_range: Column range tuple
            row_range: Row range tuple
            line_width: Grid line width
            color: Grid color

        Returns:
            List of style commands
        """
        color = color or ColorPalette.GREY

        return [
            ('GRID', (col_range[0], row_range[0]), (col_range[1], row_range[1]),
             line_width, color),
        ]

    @staticmethod
    def create_title_style(font_size: int = 16, bold: bool = True) -> List:
        """
        Create title style

        Args:
            font_size: Font size
            bold: Use bold font

        Returns:
            List of style commands
        """
        return [
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (0, 0),
             font_manager.bold_font if bold else font_manager.default_font),
            ('FONTSIZE', (0, 0), (0, 0), font_size),
            ('TOPPADDING', (0, 0), (0, 0), 6),
            ('BOTTOMPADDING', (0, 0), (0, 0), 6),
        ]


class ExcelStyleFactory:
    """Factory for creating Excel cell styles"""

    @staticmethod
    def create_header_fill() -> PatternFill:
        """Create header background fill"""
        return PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    @staticmethod
    def create_header_font(size: int = 12, bold: bool = True) -> Font:
        """Create header font"""
        return Font(color="FFFFFF", bold=bold, size=size)

    @staticmethod
    def create_cell_border() -> Border:
        """Create cell border"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @staticmethod
    def create_center_alignment() -> Alignment:
        """Create center alignment"""
        return Alignment(horizontal='center', vertical='center')

    @staticmethod
    def create_left_alignment() -> Alignment:
        """Create left alignment"""
        return Alignment(horizontal='left', vertical='center')


class DimensionHelper:
    """Helper for dimension calculations"""

    @staticmethod
    def inch(value: float) -> float:
        """Convert value to inches"""
        return value * inch

    @staticmethod
    def calculate_column_widths(total_width: float, ratios: List[float]) -> List[float]:
        """
        Calculate column widths based on ratios

        Args:
            total_width: Total available width in inches
            ratios: List of ratio values

        Returns:
            List of column widths in inches
        """
        total_ratio = sum(ratios)
        return [total_width * (ratio / total_ratio) * inch for ratio in ratios]


# Global instances
pdf_style = PDFStyleFactory()
excel_style = ExcelStyleFactory()
dimensions = DimensionHelper()
