"""Fluent helpers for building Excel worksheets."""
from typing import List, Optional, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..utils.styling import excel_style


class ExcelSheetBuilder:
    """Small fluent wrapper around an openpyxl worksheet."""

    def __init__(self, sheet: Worksheet):
        self.sheet = sheet
        self.current_row = 1
        self._header_fill = excel_style.create_header_fill()
        self._header_font = excel_style.create_header_font()
        self._cell_border = excel_style.create_cell_border()
        self._center_align = excel_style.create_center_alignment()

    def add_header_row(
        self,
        headers: List[str],
        row: Optional[int] = None
    ) -> 'ExcelSheetBuilder':
        """Add a styled header row."""
        row = row or self.current_row

        for col_num, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=row, column=col_num, value=header)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.border = self._cell_border
            cell.alignment = self._center_align

        self.current_row = row + 1
        return self

    def add_data_row(
        self,
        data: List[Any],
        row: Optional[int] = None
    ) -> 'ExcelSheetBuilder':
        """Add one bordered data row."""
        row = row or self.current_row

        for col_num, value in enumerate(data, 1):
            cell = self.sheet.cell(row=row, column=col_num, value=value)
            cell.border = self._cell_border

        self.current_row = row + 1
        return self

    def add_data_rows(
        self,
        data_rows: List[List[Any]]
    ) -> 'ExcelSheetBuilder':
        for row_data in data_rows:
            self.add_data_row(row_data)
        return self

    def add_title_row(
        self,
        title: str,
        font_size: int = 16,
        bold: bool = True
    ) -> 'ExcelSheetBuilder':
        """Add a title row at the current position."""
        cell = self.sheet.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(bold=bold, size=font_size)

        self.current_row += 1
        return self

    def add_label_value_row(
        self,
        label: str,
        value: Any,
        label_bold: bool = True
    ) -> 'ExcelSheetBuilder':
        """Add a two-column label/value row."""
        label_cell = self.sheet.cell(row=self.current_row, column=1, value=label)
        if label_bold:
            label_cell.font = Font(bold=True)

        self.sheet.cell(row=self.current_row, column=2, value=value)

        self.current_row += 1
        return self

    def skip_rows(self, count: int = 1) -> 'ExcelSheetBuilder':
        self.current_row += count
        return self

    def merge_cells(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int
    ) -> 'ExcelSheetBuilder':
        start_cell = self.sheet.cell(row=start_row, column=start_col)
        end_cell = self.sheet.cell(row=end_row, column=end_col)

        self.sheet.merge_cells(
            f"{start_cell.coordinate}:{end_cell.coordinate}"
        )
        return self

    def set_column_widths(self, widths: List[int]) -> 'ExcelSheetBuilder':
        for idx, width in enumerate(widths, start=1):
            column_letter = get_column_letter(idx)
            self.sheet.column_dimensions[column_letter].width = width
        return self

    def build(self) -> Worksheet:
        return self.sheet


class ExcelTableBuilder:
    """Collect table data before writing it to a worksheet."""

    def __init__(self):
        self.headers: List[str] = []
        self.data_rows: List[List[Any]] = []
        self.column_widths: List[int] = []

    def with_headers(self, headers: List[str]) -> 'ExcelTableBuilder':
        self.headers = headers
        return self

    def with_data(self, data_rows: List[List[Any]]) -> 'ExcelTableBuilder':
        self.data_rows = data_rows
        return self

    def with_column_widths(self, widths: List[int]) -> 'ExcelTableBuilder':
        self.column_widths = widths
        return self

    def build_to_sheet(
        self,
        sheet: Worksheet,
        start_row: int = 1
    ) -> Worksheet:
        """Write the configured table into ``sheet``."""
        builder = ExcelSheetBuilder(sheet)
        builder.current_row = start_row

        if self.headers:
            builder.add_header_row(self.headers)

        builder.add_data_rows(self.data_rows)

        if self.column_widths:
            builder.set_column_widths(self.column_widths)

        return builder.build()
