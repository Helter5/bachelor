"""Shared Excel export base class."""
from abc import ABC, abstractmethod
from io import BytesIO
from typing import Dict, Any, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from ..utils.styling import excel_style


class BaseExcelExport(ABC):
    """Base template for Excel exports.

    Subclasses load data in ``fetch_data`` and populate workbook sheets in
    ``create_sheets``. ``generate`` owns the orchestration and buffer handling.
    """

    def __init__(self):
        self.workbook: Workbook = Workbook()
        self.buffer = BytesIO()
        self.metadata: Dict[str, Any] = {}

    @abstractmethod
    def fetch_data(self) -> None:
        """Load export data into instance state."""
        pass

    @abstractmethod
    def create_sheets(self) -> None:
        """Populate workbook sheets."""
        pass

    def get_filename(self) -> str:
        """Return the default export filename."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"export_{timestamp}.xlsx"

    def validate_data(self) -> None:
        """Override when an export needs validation before rendering."""
        pass

    def apply_column_widths(self, sheet: Worksheet, widths: List[int]) -> None:
        """Apply column widths to a worksheet."""
        for idx, width in enumerate(widths, start=1):
            column_letter = get_column_letter(idx)
            sheet.column_dimensions[column_letter].width = width

    def style_header_row(
        self,
        sheet: Worksheet,
        row: int = 1,
        col_start: int = 1,
        col_end: int = None
    ) -> None:
        """Apply standard header styling to a worksheet row."""
        if col_end is None:
            col_end = sheet.max_column

        header_fill = excel_style.create_header_fill()
        header_font = excel_style.create_header_font()
        cell_border = excel_style.create_cell_border()
        center_align = excel_style.create_center_alignment()

        for col in range(col_start, col_end + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = cell_border
            cell.alignment = center_align

    def style_data_cells(
        self,
        sheet: Worksheet,
        row_start: int,
        row_end: int = None,
        col_start: int = 1,
        col_end: int = None
    ) -> None:
        """Apply standard borders to a worksheet range."""
        if row_end is None:
            row_end = sheet.max_row
        if col_end is None:
            col_end = sheet.max_column

        cell_border = excel_style.create_cell_border()

        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = cell_border

    def generate(self) -> BytesIO:
        """Generate the workbook and return a buffer positioned at the start."""
        self.fetch_data()
        self.validate_data()
        self.create_sheets()
        self.workbook.save(self.buffer)
        self.buffer.seek(0)

        return self.buffer

    def get_response_headers(self) -> Dict[str, str]:
        """Return headers for an Excel download response."""
        return {
            "Content-Disposition": f"attachment; filename={self.get_filename()}"
        }
